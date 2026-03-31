from __future__ import annotations

import math
import os
import json
from collections import deque
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
DEFAULT_WORKBOOK_PATH = Path(
    os.getenv(
        "MONITOR_WORKBOOK_PATH",
        r"D:\2 work\2兴证固收\2利率\1数据库\2中观数据库\中观数据库-维护人关梦卿\【兴证固收】中观周度数据库_20260329.xlsx",
    )
)
SNAPSHOT_PATH = ROOT / "data" / "generated" / "workbook_snapshot.json"

TRACKED_SHEETS = {
    "价格指标",
    "房地产",
    "基建",
    "工业生产",
    "耐用品消费",
    "人流物流",
    "电影",
    "出口",
    "农产品",
}
X_AXIS_TOKENS = {"日期", "报告日", "月份"}
TITLE_BLACKLIST = {
    "返回目录",
    "Wind",
    "同花顺iFinD",
    "指标名称",
    "频率",
    "单位",
    "指标ID",
    "来源",
}

_cache_lock = Lock()
_snapshot_cache: dict[str, Any] | None = None


@dataclass
class BlockBuilder:
    chart_id: str
    sheet: str
    title: str
    header_row: int
    start_col: int
    end_col: int
    headers: list[str]
    points: list[list[Any]] = field(default_factory=list)
    source: str | None = None


def _workbook_path() -> Path:
    return DEFAULT_WORKBOOK_PATH


def _as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _is_empty(value: Any) -> bool:
    return value is None or _as_text(value) == ""


def _is_number(value: Any) -> bool:
    if value is None or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        try:
            return not math.isnan(float(value))
        except Exception:  # pragma: no cover - defensive branch
            return True
    return False


def _normalize_numeric(value: Any) -> float | None:
    if not _is_number(value):
        return None
    number = float(value)
    if math.isnan(number):
        return None
    return number


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == datetime.min.time() else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"#N/A", "N/A", "-", "空"}:
            return None
        return stripped
    return value


def _title_candidate(value: Any) -> bool:
    text = _as_text(value)
    if not text or text in TITLE_BLACKLIST or text in X_AXIS_TOKENS:
        return False
    if text.startswith("数据来源") or text.startswith("根据新闻整理"):
        return False
    return True


def _find_title(history: deque[tuple[int, tuple[Any, ...]]], column_index: int) -> tuple[str | None, str | None]:
    title: str | None = None
    source: str | None = None

    for _, row_values in reversed(history):
        value = row_values[column_index - 1] if column_index - 1 < len(row_values) else None
        text = _as_text(value)
        if not text:
            continue
        if source is None and (text.startswith("数据来源") or text.startswith("根据新闻整理")):
            source = text
            continue
        if _title_candidate(value):
            title = text
            break

    return title, source


def _collect_headers(row_values: tuple[Any, ...], start_col: int) -> tuple[int, list[str]]:
    end_col = start_col
    headers: list[str] = []
    while end_col - 1 < len(row_values):
        value = row_values[end_col - 1]
        if _is_empty(value):
            break
        headers.append(_as_text(value))
        end_col += 1
    return end_col - 1, headers


def _serialize_block(block: BlockBuilder, point_limit: int | None) -> dict[str, Any]:
    rows = block.points[-point_limit:] if point_limit and point_limit > 0 else block.points
    x_values: list[Any] = []
    series_values = {header: [] for header in block.headers[1:]}

    for row in rows:
        x_values.append(_normalize_cell(row[0]))
        for idx, header in enumerate(block.headers[1:], start=1):
            value = row[idx] if idx < len(row) else None
            series_values[header].append(_normalize_numeric(value))

    series: list[dict[str, Any]] = []
    for name, values in series_values.items():
        non_null = [item for item in values if item is not None]
        if not non_null:
            continue
        series.append(
            {
                "name": name,
                "values": values,
                "last_value": non_null[-1],
                "min": min(non_null),
                "max": max(non_null),
            }
        )

    latest_values = {
        item["name"]: item["last_value"]
        for item in series
    }

    return {
        "id": block.chart_id,
        "sheet": block.sheet,
        "title": block.title,
        "source": block.source,
        "x_axis_label": block.headers[0] if block.headers else "",
        "series_count": len(series),
        "point_count": len(block.points),
        "x_values": x_values,
        "series": series,
        "latest_values": latest_values,
    }


def _parse_indicator_index(ws: Any) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    current_industry = ""
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = _as_text(row[2] if len(row) > 2 else None)
        if not name:
            continue

        industry = _as_text(row[0] if len(row) > 0 else None) or current_industry
        current_industry = industry or current_industry
        items.append(
            {
                "industry": current_industry,
                "sequence": _normalize_cell(row[1] if len(row) > 1 else None),
                "name": name,
                "indicator_id": _normalize_cell(row[3] if len(row) > 3 else None),
                "latest_status": _normalize_cell(row[4] if len(row) > 4 else None),
                "latest_window": _normalize_cell(row[5] if len(row) > 5 else None),
                "frequency": _normalize_cell(row[6] if len(row) > 6 else None),
                "direction": _normalize_cell(row[7] if len(row) > 7 else None),
                "previous_date": _normalize_cell(row[8] if len(row) > 8 else None),
                "magnitude": _normalize_cell(row[9] if len(row) > 9 else None),
                "threshold": _normalize_cell(row[10] if len(row) > 10 else None),
                "latest_date": _normalize_cell(row[11] if len(row) > 11 else None),
                "source": _normalize_cell(row[12] if len(row) > 12 else None),
            }
        )
    return items


def _parse_sheet_blocks(ws: Any) -> list[dict[str, Any]]:
    history: deque[tuple[int, tuple[Any, ...]]] = deque(maxlen=5)
    active: list[BlockBuilder] = []
    finished: list[BlockBuilder] = []

    for row_index, row_values in enumerate(ws.iter_rows(values_only=True), start=1):
        row_values = tuple(row_values)

        still_active: list[BlockBuilder] = []
        for block in active:
            segment = list(row_values[block.start_col - 1 : block.end_col])
            if len(segment) < len(block.headers):
                segment.extend([None] * (len(block.headers) - len(segment)))

            if all(_is_empty(value) for value in segment):
                finished.append(block)
                continue

            block.points.append(segment)
            still_active.append(block)
        active = still_active

        column = 1
        while column <= len(row_values):
            cell_text = _as_text(row_values[column - 1])
            if cell_text not in X_AXIS_TOKENS:
                column += 1
                continue

            end_col, headers = _collect_headers(row_values, column)
            if len(headers) < 2:
                column = max(column + 1, end_col + 1)
                continue

            title, source = _find_title(history, column)
            if not title:
                column = end_col + 1
                continue

            chart_id = f"{ws.title}-{row_index}-{column}"
            active.append(
                BlockBuilder(
                    chart_id=chart_id,
                    sheet=ws.title,
                    title=title,
                    header_row=row_index,
                    start_col=column,
                    end_col=end_col,
                    headers=headers,
                    source=source,
                )
            )
            column = end_col + 1

        history.append((row_index, row_values))

    finished.extend(active)

    serialized = []
    for block in finished:
        if len(block.points) < 5:
            continue
        preview = _serialize_block(block, point_limit=160)
        if preview["series_count"] == 0:
            continue
        preview["point_count_total"] = len(block.points)
        serialized.append(preview)

    serialized.sort(key=lambda item: item["id"])
    return serialized


def _build_snapshot() -> dict[str, Any]:
    workbook_path = _workbook_path()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    indicator_items = _parse_indicator_index(workbook["指标列表"])

    charts_by_sheet: dict[str, list[dict[str, Any]]] = {}
    charts_by_id: dict[str, dict[str, Any]] = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name not in TRACKED_SHEETS:
            continue
        chart_blocks = _parse_sheet_blocks(workbook[sheet_name])
        charts_by_sheet[sheet_name] = chart_blocks
        for chart in chart_blocks:
            charts_by_id[chart["id"]] = chart

    sheets = [
        {
            "name": sheet_name,
            "chart_count": len(charts_by_sheet.get(sheet_name, [])),
        }
        for sheet_name in workbook.sheetnames
        if sheet_name in TRACKED_SHEETS
    ]

    unique_industries = sorted({item["industry"] for item in indicator_items if item["industry"]})
    latest_dates = [item["latest_date"] for item in indicator_items if item.get("latest_date")]
    latest_updated_at = max(latest_dates) if latest_dates else datetime.fromtimestamp(workbook_path.stat().st_mtime).date().isoformat()

    workbook.close()

    return {
        "workbook_name": workbook_path.name,
        "workbook_path": str(workbook_path),
        "updated_at": latest_updated_at,
        "indicator_count": len(indicator_items),
        "industry_count": len(unique_industries),
        "sheets": sheets,
        "indicator_items": indicator_items,
        "charts_by_sheet": charts_by_sheet,
        "charts_by_id": charts_by_id,
        "mtime": workbook_path.stat().st_mtime,
    }


def _load_snapshot_file() -> dict[str, Any]:
    if not SNAPSHOT_PATH.exists():
        raise FileNotFoundError(f"Snapshot not found: {SNAPSHOT_PATH}")

    with SNAPSHOT_PATH.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    data.setdefault("mtime", None)
    return data


def _snapshot() -> dict[str, Any]:
    global _snapshot_cache

    workbook_path = _workbook_path()
    current_mtime = workbook_path.stat().st_mtime if workbook_path.exists() else None

    with _cache_lock:
        if _snapshot_cache and _snapshot_cache.get("mtime") == current_mtime:
            return _snapshot_cache

        if workbook_path.exists():
            _snapshot_cache = _build_snapshot()
        else:
            _snapshot_cache = _load_snapshot_file()
        return _snapshot_cache


def database_overview() -> dict[str, Any]:
    snapshot = _snapshot()
    highlights = snapshot["indicator_items"][:18]
    return {
        "workbook_name": snapshot["workbook_name"],
        "workbook_path": snapshot["workbook_path"],
        "updated_at": snapshot["updated_at"],
        "indicator_count": snapshot["indicator_count"],
        "industry_count": snapshot["industry_count"],
        "sheet_count": len(snapshot["sheets"]),
        "sheets": snapshot["sheets"],
        "highlights": highlights,
    }


def database_indicators(industry: str | None = None, query: str | None = None, limit: int = 200) -> dict[str, Any]:
    snapshot = _snapshot()
    items = snapshot["indicator_items"]

    if industry:
        items = [item for item in items if item["industry"] == industry]

    if query:
        keyword = query.strip().lower()
        items = [
            item
            for item in items
            if keyword in (item.get("name") or "").lower() or keyword in (item.get("latest_status") or "").lower()
        ]

    return {
        "total": len(items),
        "items": items[: max(limit, 1)],
    }


def database_charts(sheet: str | None = None, query: str | None = None, point_limit: int | None = 160) -> dict[str, Any]:
    snapshot = _snapshot()
    selected_sheet = sheet or (snapshot["sheets"][0]["name"] if snapshot["sheets"] else None)
    if not selected_sheet:
        return {"sheet": None, "charts": []}

    charts = list(snapshot["charts_by_sheet"].get(selected_sheet, []))
    if query:
        keyword = query.strip().lower()
        charts = [chart for chart in charts if keyword in chart["title"].lower()]

    if point_limit and point_limit > 0:
        charts = [
            {
                **chart,
                "x_values": chart["x_values"][-point_limit:],
                "series": [
                    {**series, "values": series["values"][-point_limit:]}
                    for series in chart["series"]
                ],
            }
            for chart in charts
        ]

    return {"sheet": selected_sheet, "charts": charts}


def database_chart(chart_id: str, point_limit: int | None = None) -> dict[str, Any]:
    snapshot = _snapshot()
    chart = snapshot["charts_by_id"].get(chart_id)
    if chart is None:
        raise KeyError(chart_id)

    if point_limit and point_limit > 0:
        return {
            **chart,
            "x_values": chart["x_values"][-point_limit:],
            "series": [
                {**series, "values": series["values"][-point_limit:]}
                for series in chart["series"]
            ],
        }
    return chart
